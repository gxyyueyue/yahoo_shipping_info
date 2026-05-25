"""
Export shipping records to a formatted Excel workbook.
"""

from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# (Excel column header, internal record key, column width)
_COLUMNS = [
    ("原始图片文件名",  "filename",       28),
    ("オークションID", "オークションID",  16),
    ("落札者ID",        "落札者ID",        18),
    ("商品名",          "商品名",          34),
    ("落札価格",        "落札価格",        12),
    ("氏名",            "氏名",            14),
    ("邮编",            "郵便番号",        12),
    ("都道府县",        "都道府県",        12),
    ("市区町村",        "市区町村",        14),
    ("详细地址",        "詳細住所",        30),
    ("配送方法",        "配送方法",        16),
    ("送料",            "送料",            10),
    ("识别状态",        "status",          12),
]

_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")

_FILL_HEADER = PatternFill("solid", fgColor="2E5BBA")
_FILL_OK     = PatternFill("solid", fgColor="EBF5EB")
_FILL_WARN   = PatternFill("solid", fgColor="FFF8E1")
_FILL_ERR    = PatternFill("solid", fgColor="FFEBEE")

# Columns that look better centre-aligned
_CENTRE_COLS = {1, 7, 13}  # 原始图片文件名, 邮编, 识别状态


def write_excel(records: List[Dict], output_path: str) -> None:
    """Write *records* to an xlsx file at *output_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "発送情報"

    _write_header(ws)
    _write_data(ws, records)

    ws.freeze_panes = "A2"
    wb.save(output_path)


def _write_header(ws) -> None:
    h_font  = Font(bold=True, color="FFFFFF", size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, (header, _, width) in enumerate(_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font      = h_font
        c.fill      = _FILL_HEADER
        c.alignment = h_align
        c.border    = _THIN
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 28


def _write_data(ws, records: List[Dict]) -> None:
    for row, rec in enumerate(records, 2):
        status = rec.get("status", "")
        if status == "エラー":
            row_fill = _FILL_ERR
        elif status == "需人工确认":
            row_fill = _FILL_WARN
        else:
            row_fill = _FILL_OK

        for col, (_, key, _) in enumerate(_COLUMNS, 1):
            c = ws.cell(row=row, column=col, value=rec.get(key, ""))
            c.border    = _THIN
            c.fill      = row_fill
            c.alignment = _CENTER if col in _CENTRE_COLS else _LEFT
