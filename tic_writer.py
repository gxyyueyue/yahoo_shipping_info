"""
Export shipping records to TIC 内贸発货单 format.

Column mapping (A–X):
  A  注文番号        ← オークションID
  B  出荷番号        ← オークションID
  C  下单时间        ← (empty)
  D  运费            ← (empty)
  E  优惠金额        ← (empty)
  F  包裹毛重(KG)    ← (empty)
  G  包裹净重(KG)    ← (empty)
  H  お届け先名称１  ← 氏名
  I  お届け先電話番号← 電話番号
  J  お届け先郵便番号← 郵便番号 (hyphen stripped)
  K  お届け先住所１  ← 都道府県
  L  お届け先住所２  ← 市区町村
  M  お届け先住所３  ← 詳細住所
  N  快递公司        ← (empty)
  O  运单号          ← (empty)
  P  配達日          ← (empty)
  Q  配達指定時間帯  ← (empty)
  R  备注            ← (empty)
  S  商品コード      ← extracted from 商品名 prefix
  T  商品名          ← 商品名 with コード prefix removed
  U  单价            ← (empty)
  V  数量            ← 1
  W  商品毛重(kg)    ← (empty)
  X  规格型号        ← (empty)
"""

import re
from datetime import date
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_HEADERS = [
    "注文番号", "出荷番号", "下单时间", "运费", "优惠金额",
    "包裹毛重(KG)", "包裹净重(KG)", "お届け先名称１", "お届け先電話番号",
    "お届け先郵便番号", "お届け先住所１", "お届け先住所２", "お届け先住所３",
    "快递公司", "运单号", "配達日", "配達指定時間帯", "备注",
    "商品コード", "商品名", "单价", "数量", "商品毛重(kg)", "规格型号",
]

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_FILL_HEADER = PatternFill("solid", fgColor="2E5BBA")
_LEFT = Alignment(horizontal="left", vertical="center")

# Stop tokens that mark the end of a product code prefix
_CODE_STOP = re.compile(r"新品|美品|★|\s")


def _strip_postal(code: str) -> str:
    """'424-0041' or '〒424-0041' → '4240041'"""
    return re.sub(r"[^\d]", "", code or "")


def _split_product_code(product_name: str) -> Tuple[str, str]:
    """Split 'ki戸11aa新品★...' into ('ki戸11aa', '新品★...').

    Rules:
      - If the name starts with a stop token (新品/美品/★/whitespace), code is empty.
      - Otherwise extract chars up to the first stop token as the code.
      - If no stop token found, code is empty and full name is returned.
    """
    if not product_name:
        return "", ""
    if _CODE_STOP.match(product_name):
        return "", product_name
    m = re.match(r"^(.+?)(?=新品|美品|★|\s)", product_name)
    if m:
        code = m.group(1)
        return code, product_name[len(code):]
    return "", product_name


def tic_output_path(shipping_xlsx_path: str) -> str:
    """Return sibling TIC file path with today's MMDD suffix."""
    mmdd = date.today().strftime("%m%d")
    parent = Path(shipping_xlsx_path).parent
    return str(parent / f"TIC内贸发货单-{mmdd}.xlsx")


def write_tic_excel(records: List[Dict], output_path: str) -> None:
    """Write *records* to a TIC 内贸発货单 xlsx at *output_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "発送情報"

    h_font  = Font(bold=True, color="FFFFFF", size=10)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    col_widths = [
        14, 14, 14, 8, 10,
        12, 12, 16, 16,
        14, 12, 16, 24,
        16, 12, 10, 16, 16,
        16, 36, 10, 6, 12, 14,
    ]

    for col, (header, width) in enumerate(zip(_HEADERS, col_widths), 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font      = h_font
        c.fill      = _FILL_HEADER
        c.alignment = h_align
        c.border    = _THIN
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28

    for row, rec in enumerate(records, 2):
        auction_id = rec.get("オークションID", "")
        product_code, product_name = _split_product_code(rec.get("商品名", ""))

        row_data = [
            auction_id,                              # A 注文番号
            auction_id,                              # B 出荷番号
            "",                                      # C 下单时间
            "",                                      # D 运费
            "",                                      # E 优惠金额
            "",                                      # F 包裹毛重
            "",                                      # G 包裹净重
            rec.get("氏名", ""),                     # H お届け先名称１
            rec.get("電話番号", "") or "080-4624-6473",  # I お届け先電話番号
            _strip_postal(rec.get("郵便番号", "")),  # J お届け先郵便番号
            rec.get("都道府県", ""),                 # K お届け先住所１
            rec.get("市区町村", ""),                 # L お届け先住所２
            rec.get("詳細住所", ""),                 # M お届け先住所３
            "",                                      # N 快递公司
            "",                                      # O 运单号
            "",                                      # P 配達日
            "",                                      # Q 配達指定時間帯
            "",                                      # R 备注
            product_code,                            # S 商品コード
            product_name,                            # T 商品名
            "",                                      # U 单价
            1,                                       # V 数量
            "",                                      # W 商品毛重
            "",                                      # X 规格型号
        ]

        for col, value in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=value)
            c.border    = _THIN
            c.alignment = _LEFT

    ws.freeze_panes = "A2"
    wb.save(output_path)
